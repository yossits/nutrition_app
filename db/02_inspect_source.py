# -*- coding: utf-8 -*-
"""
02_inspect_source.py — structure inspection for the Tzameret (Ministry of
Health) source files.

What it does:
  Scans db/source/ (or a directory passed as an argument) and, for every
  xlsx/csv file, prints: sheets, dimensions, the first 3 rows verbatim, the
  guessed header row, a profile per column (name, fill rate, numeric ratio,
  samples), the distribution of Mida values, the number of distinct CODEs,
  and the CODE overlap between files.

What it does not do:
  Writes to no DB and modifies no file. Read-only.

Run (from the repo root):
  python db/02_inspect_source.py            # reads from db/source
  python db/02_inspect_source.py <dir>      # a different directory

Output:
  db/source_report.txt  — the full report. This is the file to hand over when
  writing the loaders.
"""

import csv
import io
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is missing. Run:  pip install openpyxl")

# The Windows console is not always UTF-8; don't let Hebrew crash the script
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MAX_SAMPLE_LEN = 40          # truncate long values in the report
TOP_MIDA = 20                # how many frequent Mida values to show
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1255", "cp862")


def clip(v):
    s = "" if v is None else str(v)
    s = s.replace("\n", "\\n").replace("\r", "")
    return s if len(s) <= MAX_SAMPLE_LEN else s[: MAX_SAMPLE_LEN - 1] + "…"


def is_number(v):
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        try:
            float(v.replace(",", ""))
            return True
        except ValueError:
            return False
    return False


def looks_like(name, *needles):
    n = str(name or "").lower()
    return any(x in n for x in needles)


def guess_header_row(rows):
    """Of the first 5 rows — the one with the most non-empty text cells."""
    best, best_score = 0, -1
    for i, row in enumerate(rows[:5]):
        score = sum(1 for c in row if c is not None and not is_number(c) and str(c).strip())
        if score > best_score:
            best, best_score = i, score
    return best


def profile_table(rows, out):
    """rows: a list of rows (tuples). Prints a profile, returns (codes, midas)."""
    if not rows:
        out.write("  (empty)\n")
        return set(), []

    out.write("  First 3 rows, verbatim:\n")
    for r in rows[:3]:
        out.write("    | " + " | ".join(clip(c) for c in r[:14]))
        if len(r) > 14:
            out.write(f" | … (+{len(r)-14} columns)")
        out.write("\n")

    h = guess_header_row(rows)
    headers = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(rows[h])]
    data = rows[h + 1 :]
    out.write(f"  Guessed header row: {h + 1} · data rows: {len(data)} · columns: {len(headers)}\n\n")

    ncols = len(headers)
    non_empty = [0] * ncols
    numeric = [0] * ncols
    samples = [[] for _ in range(ncols)]
    distinct = [set() for _ in range(ncols)]

    for r in data:
        for j in range(ncols):
            v = r[j] if j < len(r) else None
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            non_empty[j] += 1
            if is_number(v):
                numeric[j] += 1
            if len(samples[j]) < 3 and clip(v) not in samples[j]:
                samples[j].append(clip(v))
            if len(distinct[j]) <= 1000:
                distinct[j].add(str(v).strip())

    out.write(f"  {'#':>3}  {'column':<28} {'fill':>9} {'numeric':>8} {'distinct':>9}  samples\n")
    for j, name in enumerate(headers):
        fill = f"{non_empty[j]}/{len(data)}" if data else "0"
        num_pct = f"{100*numeric[j]//non_empty[j]}%" if non_empty[j] else "—"
        dis = f"{len(distinct[j])}" + ("+" if len(distinct[j]) > 1000 else "")
        out.write(f"  {j:>3}  {clip(name):<28} {fill:>9} {num_pct:>8} {dis:>9}  {', '.join(samples[j])}\n")

    # Key columns
    codes = set()
    midas = []
    for j, name in enumerate(headers):
        if looks_like(name, "code", "קוד") and non_empty[j]:
            codes = distinct[j]
            out.write(f"\n  ▸ CODE column: '{name}' — {len(codes)}{'+' if len(codes) > 1000 else ''} distinct values\n")
        if looks_like(name, "mida", "מידה") and non_empty[j]:
            from collections import Counter
            cnt = Counter()
            for r in data:
                v = r[j] if j < len(r) else None
                if v is not None and str(v).strip():
                    cnt[str(v).strip()] += 1
            midas = cnt.most_common(TOP_MIDA)
            out.write(f"\n  ▸ Mida column: '{name}' — {len(cnt)} distinct values. Most frequent:\n")
            for val, c in midas:
                out.write(f"      {clip(val):<20} × {c}\n")
    return codes, midas


def read_xlsx(path, out):
    wb = load_workbook(path, read_only=True, data_only=True)
    all_codes = set()
    for ws in wb.worksheets:
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        # drop entirely empty trailing rows
        while rows and all(c is None for c in rows[-1]):
            rows.pop()
        out.write(f"\n  ── sheet: '{ws.title}' ({len(rows)} rows)\n")
        codes, _ = profile_table(rows, out)
        all_codes |= codes
    wb.close()
    return all_codes


def read_csv(path, out):
    raw = path.read_bytes()
    text, enc = None, None
    for e in CSV_ENCODINGS:
        try:
            text = raw.decode(e)
            enc = e
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        out.write("  ✗ Encoding not detected. Try opening it by hand and saving as UTF-8.\n")
        return set()
    try:
        dialect = csv.Sniffer().sniff(text[:4000], delimiters=",;\t")
        delim = dialect.delimiter
    except csv.Error:
        delim = ","
    rows = [tuple(r) for r in csv.reader(io.StringIO(text), delimiter=delim)]
    out.write(f"  encoding: {enc} · delimiter: {delim!r} · {len(rows)} rows\n")
    codes, _ = profile_table(rows, out)
    return codes


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / "source"
    if not src.is_dir():
        sys.exit(f"Directory {src} does not exist. "
                 f"Download the eight files from data.gov.il into it.")

    files = sorted(p for p in src.iterdir() if p.suffix.lower() in (".xlsx", ".xls", ".csv"))
    if not files:
        sys.exit(f"No xlsx/csv files inside {src}.")

    out = io.StringIO()
    out.write(f"Structure report — Tzameret source files\nDirectory: {src}\nFiles: {len(files)}\n")
    out.write("=" * 78 + "\n")

    codes_by_file = {}
    for p in files:
        out.write(f"\n■ {p.name}  ({p.stat().st_size:,} bytes)\n")
        try:
            if p.suffix.lower() in (".xlsx", ".xls"):
                codes_by_file[p.name] = read_xlsx(p, out)
            else:
                codes_by_file[p.name] = read_csv(p, out)
        except Exception as e:  # one broken file must not kill the report
            out.write(f"  ✗ Read error: {type(e).__name__}: {e}\n")

    # CODE overlap between files — this is what validates the linkage
    keyed = {k: v for k, v in codes_by_file.items() if v}
    if len(keyed) > 1:
        out.write("\n" + "=" * 78 + "\nCODE overlap between files:\n")
        names = list(keyed)
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                a, b = keyed[names[i]], keyed[names[j]]
                inter = len(a & b)
                out.write(f"  {names[i]} ∩ {names[j]}: {inter} shared "
                          f"(out of {len(a)} / {len(b)})\n")

    report = out.getvalue()
    dest = Path(__file__).parent / "source_report.txt"
    dest.write_text(report, encoding="utf-8")
    print(report)
    print(f"\n✔ Report saved: {dest}\n  This is the file to hand over when writing the loaders.")


if __name__ == "__main__":
    main()
